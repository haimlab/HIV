################################################
# Program to read/edit FASTA files for individual
# seasons and draw branches based on nucleotide
# sequences and create protein branched excel sheets
################################################
#Author Kallin Khan, Department of Microbiology
#Updated: 4/3/17
################################################

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl import Workbook
from ComputeStandardDevs import getVal

def AnalyzeNucleotides(Season, fasta):
    Nucs = {}
    Names = {}
    
    rowNum = 0
    rowAfterName = 0
    numElements = 0
    goodSeason = 1

    for row in fasta:
        if row[0] == '>': ##indicates the start of a new strand
            if row[-6:-1] == Season or row[-4:-1] == (Season[:2] + '|'): ##checks that the strand is in the correct season
                colNum = 1
                goodSeason = 0
                row = row[:-1]
                name = row[1:]
                Nucs[rowNum] = ''
                Names[rowNum] = name
                numElements = 0
                rowAfterName = 1
                rowNum += 1
                print("Creating Row: " + str(rowNum))
            else:
                goodSeason = 1
                print("Removing: " + row)
        elif goodSeason == 0: ##if the strand is in the correct season, this loop will execute
            if rowAfterName == 1: ##the first row after a name may often start incorrectly (dashes, extra elements)
                rowAfterName = 0
                while row[:3] != "ATG" : ##loop to eliminate anything before the correct start of the strand
                     row = row[1:]
            for el in row:
                if el != "-" and el != '\n' and numElements < 1699: ##stops after the correct number of elements
                    numElements += 1
                    colNum += 1
                    Nucs[rowNum-1] += el

    return (rowNum+1,Nucs,Names)

def CreateExcelSheetsForProteins(book, Season, correctSequence, fasta, rowNumbers, nSites, offSet):
    sheet1 = book.active
    if offSet == 2:
        sheet1.title = "Protein Sequence"
        sheet2 = book.create_sheet("Branches-Pro")
    else:
        sheet2 = book["Branches-Pro"]
    rowNum = offSet-1
    rowAfterName = 0
    numElements = 0
    goodSeason = 1
    Xs = {""}

    for i in range(1,551):
        cell = sheet2.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')
        cell = sheet1.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')

    for row in fasta:
        if row[0] == '>': ##indicates the start of a new strand
            if row[-6:-1] == Season or row[-4:-1] == (Season[:2] + '|'): ##checks that the strand is in the correct season
                rowNum += 1
                colNum = 1
                goodSeason = 0
                row = row[:-1]
                sheet1.cell(row=rowNum, column=1, value=row[1:]) ##adds the strand name to the first column of the spreadsheet
                sheet2.cell(row=rowNumbers[rowNum-offSet], column=1, value=row[1:]) ##adds the strand name to the first column of the spreadsheet
                numElements = 0
                rowAfterName = 1
                print("Creating: " + str(rowNum))
            else:
                goodSeason = 1
                print("Removing: " + row)
        elif goodSeason == 0: ##if the strand is in the correct season, this loop will execute

            for el in row:
                if el != "-" and el != '\n' and numElements < 1699: ##stops after the correct number of elements
                    color = False
                    if getVal(el) == -100:
                        el = correctSequence[colNum-1]
                        color = True
                        Xs.add(sheet1['A' + str(rowNum)].value)
                    if colNum+1 in nSites:
                        colNum += 1
                        numElements += 1
                        cell1 = sheet1.cell(row=rowNum, column=colNum, value='Z')
                        cell1.alignment = Alignment(horizontal='center')
                        cell2 = sheet2.cell(row=rowNumbers[rowNum-offSet], column=colNum, value='Z')
                        cell2.alignment = Alignment(horizontal='center')
                    else:
                        numElements += 1
                        colNum += 1
                        cell1 = sheet1.cell(row=rowNum, column=colNum, value=el)
                        cell1.alignment = Alignment(horizontal='center')                    
                        cell2 = sheet2.cell(row=rowNumbers[rowNum-offSet], column=colNum, value=el)
                        cell2.alignment = Alignment(horizontal='center')
                        if el != correctSequence[colNum-2] or color:
                            if color:
                                cell1.fill = PatternFill('solid', fgColor="FF0000")
                                cell2.fill = PatternFill('solid', fgColor="FF0000")
                            else:
                                cell1.fill = PatternFill('solid', fgColor="DDDDDD")
                                cell2.fill = PatternFill('solid', fgColor="DDDDDD")                            
            
    return (rowNum, Xs)

def findZSites(correctSeq):
    oneAgo = False
    twoAgo = False
    nSites = {0}
    loc = 0
    for l in correctSeq:
        if twoAgo:
            if l == 'S' or l == 'T':
                nSites.add(loc)
            twoAgo = False
        if oneAgo:
            twoAgo = True
            oneAgo = False
        if l == 'N':
            oneAgo = True
        loc += 1

    nSites.remove(0)
    return nSites

def saveExcel(book, Season, State, years):
    print('Saving...')
    book.save("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\"+State+"\\"+ years+"\\Protein Data" + "-" + State + "-" + Season + ".xlsx")
