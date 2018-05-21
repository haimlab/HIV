################################################
# Program to show the duplicates in an Excel
# spreadsheet of branches and delete them,
# creating a new sheet without duplicates.
################################################
#Author Kallin Khan, Department of Microbiology
#Updated: 1/31/17
################################################

import sys 
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors

def DeleteDuplicates(book, numRows, correctSequence, numBranches,file):
    sheet1 = book["Protein Sequence"]
    sheet2 = book.create_sheet("Pro-NoDup")

    rows = {''}
    rowNum = 2

    for i in range(1,551):
        cell = sheet2.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')
    
    for length in range(2,numRows+1+numBranches):
        row1 = ''
        
        compName1 = 'B' + str(length)
        compName2 = 'UU' + str(length)
        rowComp = sheet1[compName1:compName2] ##sequence of the row for comparison to others

        if type(rowComp[0][0].value) is not str:   ##will skip the blank lines created by branch gaps
            sheet2.cell(row=rowNum, column=1, value="")
            rowNum += 1
            continue
        
        for val in rowComp:
            for  val1 in val:
                row1 += val1.value ##a string created from the above sequence
        print('Checking Row ' + str(length-1))
        if row1 not in rows: ##checks duplication
            rows.add(row1)
            rowName = 'A' + str(length)
            sheet2.cell(row=rowNum, column=1, value=sheet1[rowName].value)
            file.write('>' + sheet1[rowName].value + '\n' + row1 + '\n')
            colNum = 2
            for el in row1:
                if el == 'X':
                    el = correctSequence[colNum-2]
                cell = sheet2.cell(row=rowNum,column=colNum,value=el)
                cell.alignment = Alignment(horizontal='center')
                if el != correctSequence[colNum-2] and el != 'Z':
                    cell.fill = PatternFill('solid', fgColor="DDDDDD")
                colNum += 1
            
            rowNum += 1
           
        else:
            print('Duplicate')
    file.close()
