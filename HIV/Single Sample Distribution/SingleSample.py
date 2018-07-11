#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul  2 15:41:36 2018

@author: Han
"""

import xlrd
import openpyxl

# this program calculates every single sample's amino acids distribution 
#Steps
#1. change the Input directory, don't delete r 
#2. change the Output directory, don't delete r. When you copy the address, keep in mind there's also a slash at the end.
#3. Make a name of the Output file                 
#4. change the Position 


Input = r"/Users/Han/Documents/Haim Lab(2018 summer)/Single Sample Distribution/B Europe Single TP 2_more Envs_W_T-R_PAIRS.xlsx"#AllEnvs_B_NA.xlsx"
Output = r"/Users/Han/Documents/Haim Lab(2018 summer)/Single Sample Distribution/"
OutputName = "EuropePosition448.xlsx"
Position = [448]
##################################################################
##################################################################
##################################################################
##################################################################
##################################################################
#Code:


workbook = xlrd.open_workbook(Input)
sheet = workbook.sheet_by_index(0)
nRows = sheet.nrows
nCols = sheet.ncols

#get data of row y by index
def GetRowIndexData(y):
    row = []
    for col in range (nCols):
        row.append(sheet.cell_value((y-1),col))
    return row

#get data of column x by index
def GetColIndexData(x):
    col = []
    for row in range (nRows):
        col.append(sheet.cell_value(row,(x-1)))
    return col

#get data of column x by title of first row 
def GetColData(x): 
    
    for i in GetRowIndexData(1):
        if i == x : 
            a = GetColIndexData(GetRowIndexData(1).index(i)+1)
            del a[0]
            return a
#get data at a specific location
def GetDataAt(y,x):
    return (sheet.cell_value((y-1),(x-1)))
# get a list contains all parients
def getAllPatient():
    result = []
    for i in GetColData("Patient"):
        if i not in result:
            result.append(i)  
    return result
# get a list contains all years' index
def getAllYearsIndex():
    result = []
    for i in AllPatient:
        count = 0
        for j in GetColData("Patient"):
            if i == j:
                result.append(count)
                break
            else:
                count = count + 1                
    return result 
AllPatient = getAllPatient()
YearIndex = getAllYearsIndex()

# pair year and patient
def getPatientAndYear():
    result = []
    for i in AllPatient:
        inside = []
        inside.append(i)
        result.append(inside)
    for j in range(len(YearIndex)):
        inin = []
        inin.append(GetDataAt(YearIndex[j]+2,6))
        inin.append(GetDataAt(YearIndex[j]+2,6))
        result[j].append(inin)
    return result
        

PatientAndYear = getPatientAndYear()#looks like this: [["MM8",[2001,2001]]]

# NEED TO BE IMPROVED!!! This part was designed for multiple patents selection. The input was not 
# only positions, but also patient's name and year. However it turned out we need to show all 
# patients' distribution, so this part should be changed. 
def NarrowDownToPatientAndYear():
    resultresult = []
    for m in Position:
        pos = GetRowIndexData(1).index(m)+1
        result = []
        for i in PatientAndYear: #[["CW002",[1997,1997]],["CW010",[1997,1997]]]
            part = []
            #print (i)
            for j in range(len(GetColData("Patient"))):
            #print (j)
                if GetColData("Patient")[j] == i[0] and GetDataAt(j+2,6) == i[1][0]:
                #print(GetDataAt(j+1,pos))
                    part.append(GetDataAt(j+2,pos))
            result.append(part)           
        resultresult.append(result)            
    return resultresult      

test1 = NarrowDownToPatientAndYear()

#cal the sequence number of Patient in a certain year, for x in PatientAndYear
def sequenceNum(x):
    count = 0
    for j in range(len(GetColData("Patient"))):
        if GetColData("Patient")[j] == x[0] and GetDataAt(j+2,6) >= x[1][0] and GetDataAt(j+2,6) <= x[1][1]:
                count = count + 1
    return count

# All amino acids
AA = ['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y','Z']
#count amino acids for each patient
numberList = []
def countAA():
    for i in test1:       
        numnumList = []
        for q in i:
            numList = []
            for aa in AA:
                num = 0
                for h in q:
                    if aa == h:
                        num = num + 1
                numList.append(num)
            numnumList.append(numList)             
        numberList.append(numnumList)
    return numberList
#numberList
numberList = countAA()
#sum
SumOfSeq = []
for i in PatientAndYear:
    SumOfSeq.append(sequenceNum(i))
#percentage = str(numberList / Sum * 100) + "%"
#write numberList into percentList
def percentage():
    for i in range(len(numberList)):
        for j in range(len(numberList[i])):
            for p in range(len(numberList[i][j])):
                numberList[i][j][p] = str(numberList[i][j][p]/SumOfSeq[j] * 100)+"%"
    return          
percentage()
#write into excel file

def writeFile():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'
#Loop to set the value of each cell
    rCount = 1
    cCount = 1
    subCount = 2
    for i in Position:
        rCount = 1
        sheet.cell(row=rCount, column=cCount).value = i
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Patient"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Year"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "# of Sequence"
        rCount = rCount + 1          
        for j in range(len(AA)):
            sheet.cell(row=rCount, column=cCount).value = AA[j]
            rCount = rCount +1       
        cCount = cCount + 1       
        rCount = 2
        for k in PatientAndYear:
            rCount = 2
            sheet.cell(row=rCount, column=cCount).value = k[0]
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = str(k[1][0])
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = sequenceNum(k)    
            cCount = cCount+1       
        for f in range(len(numberList[Position.index(i)])):
                for g in range(len(numberList[Position.index(i)][f])):
                    sheet.cell(row=g+5, column=subCount).value = numberList[Position.index(i)][f][g]
                subCount = subCount+1 
        subCount = subCount+4
        cCount = cCount+3
    #print(numberList)
    wb.save(Output+OutputName)
writeFile()

            
                
            
    




