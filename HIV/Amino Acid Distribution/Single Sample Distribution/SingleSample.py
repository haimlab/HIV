#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul  2 15:41:36 2018

@author: Changze Han
"""

import xlrd
import openpyxl

# this program calculates every single sample's amino acids distribution 
#Steps
#1. change the Input directory, don't delete r 
#2. change the Output directory, don't delete r. When you copy the address, keep in mind there's also a slash at the end.
#3. Make a name of the Output file                 
#4. change the Position 


Input =r"/Users/Han/Documents/Haim Lab(2018 summer)/8.1.18 Clade C Single Samples/Clade C input/Eastern Central_Africa_2(Use in Code).xlsx"#3B Europe Single TP 2_more Envs_W_T-R_PAIRS.xlsx"#AllEnvs_B_NA.xlsx"
Output = r"/Users/Han/Documents/Haim Lab(2018 summer)/8.1.18 Clade C Single Samples/"
OutputName = "C_ECA2_295.xlsx"#"7.31.18Single C_ECA2_667.xlsx"
Position = 295
##################################################################
##################################################################
##################################################################
##################################################################
##################################################################
#Code:


workbook = xlrd.open_workbook(Input)
sheet = workbook.sheet_by_index(0)
nRows = sheet.nrows
nCols = sheet.ncols

#get data of row y by index
def GetRowIndexData(y):
    row = []
    for col in range (nCols):
        row.append(sheet.cell_value((y-1),col))
    return row

#get data of column x by index
def GetColIndexData(x):
    col = []
    for row in range (nRows):
        col.append(sheet.cell_value(row,(x-1)))
    return col

#get data of column x by title of first row 
def GetColData(x): 
    Row1 = GetRowIndexData(1)
    for i in Row1:
        if i == x : 
            a = GetColIndexData(Row1.index(i)+1)
            del a[0]
            return a
#get data at a specific location
def GetDataAt(y,x):
    return (sheet.cell_value((y-1),(x-1)))
# get a list contains all parients' name
Patient = GetColData("Patient")######################### list of Patient column 
Year = GetColData("Year")######################### list of Year column 
def getAllPatient():
    result = []
    
    for i in Patient:
        if i not in result:
            result.append(i)  
    return result
# get a list contains all years' index
'''def getAllYearsIndex():
    result = []
    for i in AllPatient:
        count = 0
        for j in GetColData("Patient"):
            if i == j:
                result.append(count)
                break
            else:
                count = count + 1                
    return result '''
def GetYearList():
    result = []
    for i in AllPatient:
        result.append(Year[Patient.index(i)])
    return result
        
AllPatient = getAllPatient()######################### ['LAI',20044616.0,'CW002','CW010','CW012','CW048','MM4']
YearList = GetYearList() ######################### looks like[0, 1, 2, 12, 29, 48], so that can get year list [1983,2004,1997...]
def GetAAList(): # get each patient's amino acids lists
    final = []
    xOfPosition = GetRowIndexData(1).index(Position)+1 # x axis index of position in excel file
    for i in AllPatient:
        result = []
        for j in range(len(Patient)):
            if Patient[j] == i:
                result.append(GetDataAt(j+2,xOfPosition))
        final.append(result)
    return final
AAList = GetAAList() ######################### a list contains each patient's amino acids
AA = ['Z','N','T','S','D','E','K','R','H','Y','Q','I','L','V','A','C','F','G','M','P','W'] #['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y','Z']
def CountAAList(): # count the distribution of each patient's AA
    final = []
    for i in AAList:
        result = []
        for j in AA:
            count = 0
            for k in i:
                if k == j:
                    count = count +1
            result.append(count)
        final.append(result)
    return final
AADistribution = CountAAList()######################### the distribution of each patient's AA List

def GetNumOfSequence():
    result = []
    for i in AllPatient:
        count = 0
        for j in Patient:
            if j == i:
                count = count + 1
        result.append(count)
    return result

NumOfSequence = GetNumOfSequence()######################### list contains each patient's sequence number

Country = GetColData("Country")
def GetCountryList():
    result = []
    for i in AllPatient:
        result.append(Country[Patient.index(i)])
    return result
CountryList = GetCountryList()######################### 

def writeFile():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'
    
    sheet.cell(row=1, column=1).value = Position
    sheet.cell(row=2, column=1).value = "Patient"
    sheet.cell(row=3, column=1).value = "Year"
    for i in range(len(AA)): 
        sheet.cell(row=i+4, column=1).value = AA[i]
    for i in range(len(AllPatient)):
        sheet.cell(row=2, column=i+2).value = AllPatient[i]
    for i in range(len(YearList)):
        sheet.cell(row=3, column=i+2).value = YearList[i]
    for i in range(len(AADistribution)):
        for j in range(len(AADistribution[i])):
            sheet.cell(row=j+4, column=i+2).value = AADistribution[i][j]
    sheet.cell(row=25, column=1).value = "#OfSequence"
    for i in range(len(NumOfSequence)):
        sheet.cell(row=25, column=i+2).value = NumOfSequence[i]
        
    for i in range(len(CountryList)):
        sheet.cell(row=26, column=i+2).value = CountryList[i]
    wb.save(Output+OutputName)
writeFile()
'''    
#Loop to set the value of each cell
    rCount = 1
    cCount = 1
    subCount = 2
    for i in Position:
        rCount = 1
        sheet.cell(row=rCount, column=cCount).value = i
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Patient"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Year"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "# of Sequence"
        rCount = rCount + 1          
        for j in range(len(AA)):
            sheet.cell(row=rCount, column=cCount).value = AA[j]
            rCount = rCount +1       
        cCount = cCount + 1       
        rCount = 2
        for k in PatientAndYear:
            rCount = 2
            sheet.cell(row=rCount, column=cCount).value = k[0]
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = str(k[1][0])
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = sequenceNum(k)    
            cCount = cCount+1       
        for f in range(len(numberList[Position.index(i)])):
                for g in range(len(numberList[Position.index(i)][f])):
                    sheet.cell(row=g+5, column=subCount).value = numberList[Position.index(i)][f][g]
                subCount = subCount+1 
        subCount = subCount+4
        cCount = cCount+3
    #print(numberList)
    wb.save(Output+OutputName)
writeFile()


# pair year and patient
def getPatientAndYear():
    result = []
    for i in AllPatient:
        inside = []
        inside.append(i)
        result.append(inside)
    for j in range(len(YearIndex)):
        inin = []
        inin.append(GetDataAt(YearIndex[j]+2,6))
        inin.append(GetDataAt(YearIndex[j]+2,6))
        result[j].append(inin)
    return result
        

PatientAndYear = getPatientAndYear()#looks like this: [["MM8",[2001,2001]]]

# NEED TO BE IMPROVED!!! This part was designed for multiple patents selection. The input was not 
# only positions, but also patient's name and year. However it turned out we need to show all 
# patients' distribution, so this part should be changed. Otherwise it's gonna take very long running time
def NarrowDownToPatientAndYear():
    resultresult = []
    pos = GetRowIndexData(1).index(Position[0])+1 # pos is the col index of Position[0]
    result = []
    for i in PatientAndYear: #[["CW002",[1997,1997]],["CW010",[1997,1997]]]
        part = []
            #print (i)
        for j in range(len(GetColData("Patient"))):
            #print (j)
            if GetColData("Patient")[j] == i[0]:# and GetDataAt(j+2,6) == i[1][0]:
                #print(GetDataAt(j+1,pos))
                part.append(GetDataAt(j+2,pos))
        result.append(part)           
    resultresult.append(result)            
    return resultresult      
test1 = NarrowDownToPatientAndYear()

#cal the sequence number of Patient in a certain year, for x in PatientAndYear
def sequenceNum(x):
    count = 0
    for j in range(len(GetColData("Patient"))):
        if GetColData("Patient")[j] == x[0] and GetDataAt(j+2,6) >= x[1][0] and GetDataAt(j+2,6) <= x[1][1]:
                count = count + 1
    return count

# All amino acids
AA = ['Z','N','T','S','D','E','K','R','H','Y','Q','I','L','V','A','C','F','G','M','P','W'] #['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y','Z']
#count amino acids for each patient
numberList = []
def countAA():
    for i in test1:       
        numnumList = []
        for q in i:
            numList = []
            for aa in AA:
                num = 0
                for h in q:
                    if aa == h:
                        num = num + 1
                numList.append(num)
            numnumList.append(numList)             
        numberList.append(numnumList)
    return numberList
#numberList
numberList = countAA()
#sum
SumOfSeq = []
for i in PatientAndYear:
    SumOfSeq.append(sequenceNum(i))

#percentage = str(numberList / Sum * 100) + "%"
#write numberList into percentList
def percentage():
    for i in range(len(numberList)):
        for j in range(len(numberList[i])):
            for p in range(len(numberList[i][j])):
                numberList[i][j][p] = str(numberList[i][j][p]/SumOfSeq[j] * 100)
    return          
percentage()
#write into excel file

def writeFile():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'
#Loop to set the value of each cell
    rCount = 1
    cCount = 1
    subCount = 2
    for i in Position:
        rCount = 1
        sheet.cell(row=rCount, column=cCount).value = i
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Patient"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "Year"
        rCount = rCount+1
        sheet.cell(row=rCount, column=cCount).value = "# of Sequence"
        rCount = rCount + 1          
        for j in range(len(AA)):
            sheet.cell(row=rCount, column=cCount).value = AA[j]
            rCount = rCount +1       
        cCount = cCount + 1       
        rCount = 2
        for k in PatientAndYear:
            rCount = 2
            sheet.cell(row=rCount, column=cCount).value = k[0]
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = str(k[1][0])
            rCount = rCount + 1
            sheet.cell(row=rCount, column=cCount).value = sequenceNum(k)    
            cCount = cCount+1       
        for f in range(len(numberList[Position.index(i)])):
                for g in range(len(numberList[Position.index(i)][f])):
                    sheet.cell(row=g+5, column=subCount).value = numberList[Position.index(i)][f][g]
                subCount = subCount+1 
        subCount = subCount+4
        cCount = cCount+3
    #print(numberList)
    wb.save(Output+OutputName)
writeFile()
'''

            
                
            
    




