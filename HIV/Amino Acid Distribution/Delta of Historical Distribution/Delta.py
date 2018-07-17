#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 11 14:38:59 2018

@author: Han
"""

# this program calculate the amount of change (delta) based on the historical distribution file

# Steps:
# 1. Change the Input directory, do not delete r
# 2. Change the Output directory, don't delete r. When you copy the address, keep in mind there's also a slash at the end.
# 3. Make a name of the Output file 
# 4. Change the sheet number of input Excel file 

Input = r"/Users/Han/Documents/Haim Lab(2018 summer)/delta/Updated Year Range's Delta/distribution Data/448distribution.xlsx"
Output = r"/Users/Han/Documents/Haim Lab(2018 summer)/delta/Updated Year Range's Delta/"
OutputName = "Position448Delta.xlsx"
sheetNumber = 0

#################################################
#################################################
#################################################
#################################################

# Code: 
import xlrd
import openpyxl


workbook = xlrd.open_workbook(Input)
sheet = workbook.sheet_by_index(sheetNumber)
nRows = sheet.nrows
nCols = sheet.ncols

#get data of row y by index
def GetRowIndexData(y):
    row = []
    for col in range (nCols):
        row.append(sheet.cell_value((y-1),col))
    return row

#get data of column x by index
def GetColIndexData(x):
    col = []
    for row in range (nRows):
        col.append(sheet.cell_value(row,(x-1)))
    return col

#get data of column x by title of first row, Eg: GetColData("[1979, 1986]")
def GetColData(x): 
    
    for i in GetRowIndexData(1):
        if i == x : 
            a = GetColIndexData(GetRowIndexData(1).index(i)+1)
            del a[0]
            return a
#get data at a specific location
def GetDataAt(y,x):
    return (sheet.cell_value((y-1),(x-1)))

#generate a list contains all time periods' column data
def getAllPeriodsCol():
    result = []
    for i in range(nCols+1)[2:]:
        result.append(GetColIndexData(i)[1:-1])
    return result

AllPeriodsData = getAllPeriodsCol()
#calculate delta, only 5 combinations, P1 & P2,P2 & P3,P3 & P4,P4 & P5,P5 & P6
def FiveCombinationsDelta():
    finalResult = []
    firstFive = AllPeriodsData[:-1]
    secondFive = AllPeriodsData[1:]
    for i in range(len(firstFive)):
        result = []
        for j in range(len(firstFive[i])):
            result.append(secondFive[i][j] - firstFive[i][j])
        finalResult.append(result)
    return finalResult

FiveResults = FiveCombinationsDelta()
#calculate delta, all combinations eg: P1 & P2,P1 & P3,P1 & P4,P1 & P5.....
def AllCombinationsDelta():
    finalResult = []
    firstFive = AllPeriodsData[:-1]
    secondFive = AllPeriodsData[1:]
    for i in firstFive:
        for j in secondFive:
            result = []
            for k in range(len(i)):
                result.append(j[k] - i[k])
            finalResult.append(result)
        secondFive = secondFive[1:]
    return finalResult

AllResults = AllCombinationsDelta()

# write into file
def writeFile():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'
    rCount = 8
    cCount = 1
    Periods = ["P1","P2","P3","P4","P5","P6"]
    for j in range(len((GetRowIndexData(1)[1:]))):#add all periods at the top of the file
        sheet.cell(row=1+j, column=cCount).value = Periods[j]+GetRowIndexData(1)[1:][j]
    AA = GetColIndexData(1)[1:]
    for i in range(21): # add first row
        sheet.cell(row=rCount, column=cCount+i+1).value = AA[i]
    sheet.cell(row=rCount, column=cCount).value = GetRowIndexData(1)[0]# add position 
    for k in range(len(Periods)-1):# add first column
        sheet.cell(row=rCount+k+1, column=cCount).value = Periods[k]+Periods[k+1]
    rCount = rCount+1
    
    for i in range(len(FiveResults)): # add five combination results to file
        for j in range(len(FiveResults[i])):
            sheet.cell(row=rCount, column=cCount+j+1).value = FiveResults[i][j]
        rCount = rCount+1
    # add all combination results to file
    rCount = rCount+1
    for i in range(21): # add first row
        sheet.cell(row=rCount, column=cCount+i+1).value = AA[i]
    # add first column
    FirstFive = Periods[:-1]
    SecondFive = Periods[1:]
    subRC = rCount
    for i in FirstFive:#range(len(FirstFive)):
        for j in SecondFive:#range(len(SecondFive)):
            sheet.cell(row=subRC+1, column=1).value = i+j#FirstFive[i]+SecondFive[j]  
            subRC = subRC + 1
        SecondFive = SecondFive[1:]
    #add position
    sheet.cell(row=rCount, column=cCount).value = GetRowIndexData(1)[0]
    # add data
    for i in range(len(AllResults)):
        for j in range(len(AllResults[i])):
            sheet.cell(row=rCount+1, column=cCount+j+1).value = AllResults[i][j]
        rCount = rCount+1
    wb.save(Output+OutputName)
writeFile()
    
    
    
    
    
    
    
    
    