################################################
# Program to read/edit FASTA files for several
# seasons and compare difference across states
################################################
#Author Kallin Khan, Department of Microbiology
#Updated: 3/9/17
################################################
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl import Workbook
import collections

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl import Workbook
import math
import collections
import copy
import csv

def computeMostCommon(fasta, Season):
    newFasta = []
    colEls = ['']*566
    rowNum = 1
    for row in fasta:
        newFasta.append(row)
        if row[0] == '>': ##indicates the start of a new strand
            if row[-6:-1] == Season or row[-4:-1] == (Season[:2] + '|'): ##checks that the strand is in the correct season
                rowNum += 1
                colNum = 1
                goodSeason = 0
                numElements = 0
                rowAfterName = 1
            else:
                goodSeason = 1
                print(row)
        elif goodSeason == 0: ##if the strand is in the correct season, this loop will execute
            if rowAfterName == 1: ##the first row after a name may often start incorrectly (dashes, extra elements)
                rowAfterName = 0 
                helper = {}
                helperCount = 0
                while row[:3] != "MKT" : ##loop to eliminate anything before the correct start of the strand
                    helper[helperCount] = row[0]
                    helperCount += 1
                    row = row[1:]
                    if row[:3] == "KTI":
                        fix = helper[0] + row
                        row = fix
                        break
                    elif row[:3] == "TII":
                        fix = helper[0] + helper[1] + row
                        row = fix
                        break
                    elif row[:3] == "IIA":
                        fix = helper[0] + helper[1] + helper[2] + row
                        row = fix
                        break                 
            for el in row:
                if el != "-" and el != '\n':
                    numElements += 1
                    colNum += 1
                    colEls[colNum-2] += el
        
    correctSequence = ""
    for el in colEls:
        correctSequence += ((collections.Counter(el).most_common(1)[0])[0])

    return (correctSequence, newFasta)
    

def createExcelSheetPro(sheet1, sheet2, Season, correctSequence, fasta, rowNumbers, nSites, offSet):

    rowNum = offSet-1
    rowAfterName = 0
    numElements = 0
    goodSeason = 1
    Xs = {""}
    
    for i in range(1,551):
        cell = sheet2.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')
        cell = sheet1.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')

    for row in fasta:
        if row[0] == '>': ##indicates the start of a new strand
            if row[-6:-1] == Season or row[-4:-1] == (Season[:2] + '|'): ##checks that the strand is in the correct season
                rowNum += 1
                colNum = 1
                goodSeason = 0
                row = row[:-1]
                sheet1.cell(row=rowNum, column=1, value=row[1:]) ##adds the strand name to the first column of the spreadsheet
                sheet2.cell(row=rowNumbers[rowNum-offSet], column=1, value=row[1:]) ##adds the strand name to the first column of the spreadsheet
                numElements = 0
                rowAfterName = 1
                print("Creating: " + str(rowNum))
            else:
                goodSeason = 1
                print("Removing: " + row)
        elif goodSeason == 0: ##if the strand is in the correct season, this loop will execute
            if rowAfterName == 1: ##the first row after a name may often start incorrectly (dashes, extra elements)
                rowAfterName = 0
                helper = {}
                helperCount = 0
                while row[:3] != "MKT" : ##loop to eliminate anything before the correct start of the strand
                    helper[helperCount] = row[0]
                    helperCount += 1
                    row = row[1:]
                    if row[:3] == "KTI":
                        fix = helper[0] + row
                        row = fix
                        break
                    elif row[:3] == "TII":
                        fix = helper[0] + helper[1] + row
                        row = fix
                        break
                    elif row[:3] == "IIA":
                        fix = helper[0] + helper[1] + helper[2] + row
                        row = fix
                        break  
            for el in row:
                if el != "-" and el != '\n' and numElements < 1699: ##stops after the correct number of elements
                    color = False
                    if getVal(el) == -100:
                        el = correctSequence[colNum-1]
                        color = True
                        Xs.add(sheet1['A' + str(rowNum)].value)
                    if colNum+1 in nSites:
                        colNum += 1
                        numElements += 1
                        cell1 = sheet1.cell(row=rowNum, column=colNum, value='Z')
                        cell1.alignment = Alignment(horizontal='center')
                        cell2 = sheet2.cell(row=rowNumbers[rowNum-offSet], column=colNum, value='Z')
                        cell2.alignment = Alignment(horizontal='center')
                    else:
                        numElements += 1
                        colNum += 1
                        cell1 = sheet1.cell(row=rowNum, column=colNum, value=el)
                        cell1.alignment = Alignment(horizontal='center')                    
                        cell2 = sheet2.cell(row=rowNumbers[rowNum-offSet], column=colNum, value=el)
                        cell2.alignment = Alignment(horizontal='center')
                        if el != correctSequence[colNum-2] or color:
                            if color:
                                cell1.fill = PatternFill('solid', fgColor="FF0000")
                                cell2.fill = PatternFill('solid', fgColor="FF0000")
                            else:
                                cell1.fill = PatternFill('solid', fgColor="DDDDDD")
                                cell2.fill = PatternFill('solid', fgColor="DDDDDD")                            
            

  
    return (rowNum, Xs)

def createExcelSheetNuc(Season, fasta):
    #sheet1 = book.active
    #sheet1.title = "Nucleotide Sequence"

    Nucs = {}
    Names = {}
    
    rowNum = 1
    rowAfterName = 0
    numElements = 0
    goodSeason = 1

    for row in fasta:
        if row[0] == '>': ##indicates the start of a new strand
            if row[-6:-1] == Season or row[-4:-1] == (Season[:2] + '|'): ##checks that the strand is in the correct season
                rowNum += 1
                colNum = 1
                goodSeason = 0
                row = row[:-1]
                name = row[1:]
                #sheet1.cell(row=rowNum, column=1, value=name) ##adds the strand name to the first column of the spreadsheet
                Nucs[rowNum-2] = ''
                Names[rowNum-2] = name
                numElements = 0
                rowAfterName = 1
                print("Creating Row: " + str(rowNum))
            else:
                goodSeason = 1
                print("Removing: " + row)
        elif goodSeason == 0: ##if the strand is in the correct season, this loop will execute
            if rowAfterName == 1: ##the first row after a name may often start incorrectly (dashes, extra elements)
                rowAfterName = 0
                while row[:3] != "ATG" : ##loop to eliminate anything before the correct start of the strand
                     row = row[1:]
            for el in row:
                if el != "-" and el != '\n' and numElements < 1699: ##stops after the correct number of elements
                    numElements += 1
                    colNum += 1
                    #cell = sheet1.cell(row=rowNum, column=colNum, value=el)
                    #cell.alignment = Alignment(horizontal='center')
                    Nucs[rowNum-2] += el

    return (rowNum,Nucs,Names)

def deleteDuplicates(book, numRows, correctSequence, numBranches):
    sheet1 = book["Branches-Pro"]
    sheet2 = book.create_sheet("Branches-Pro-NoProtein-Dup")

    rows = {''}
    rowNum = 2

    for i in range(1,551):
        cell = sheet2.cell(row=1, column=17+i, value=i)
        cell.alignment = Alignment(horizontal='center')
    
    for length in range(2,numRows+1+numBranches):
        row1 = ''
        
        compName1 = 'B' + str(length)
        compName2 = 'UU' + str(length)
        rowComp = sheet1[compName1:compName2] ##sequence of the row for comparison to others

        if type(rowComp[0][0].value) is not str:   ##will skip the blank lines created by branch gaps
            sheet2.cell(row=rowNum, column=1, value="")
            rowNum += 1
            continue
        
        for val in rowComp:
            for  val1 in val:
                row1 += val1.value ##a string created from the above sequence
        print('Checking Row ' + str(length-1))
        if row1 not in rows: ##checks duplication
            rows.add(row1)
            rowName = 'A' + str(length)
            sheet2.cell(row=rowNum, column=1, value=sheet1[rowName].value)
            colNum = 2
            for el in row1:
                if el == 'X':
                    el = correctSequence[colNum-2]
                cell = sheet2.cell(row=rowNum,column=colNum,value=el)
                cell.alignment = Alignment(horizontal='center')
                if el != correctSequence[colNum-2] and el != 'Z':
                    cell.fill = PatternFill('solid', fgColor="DDDDDD")
                colNum += 1
            
            rowNum += 1
           
        else:
            print('Duplicate')
        
                
        
def compBranch(Reader,numberOfRows, offSet):
    num = numberOfRows-1 #int(input("Number of cases: "))

    Matrix = [[20.0 for x in range(num)] for y in range(num)]
    Names = ['' for x in range(num)]
    Branch = [-1 for x in range(num)]

    i = 0
    rowNum = 0

    for row in Reader: ##reads each row of the csv
        colNum = 0
        t = (', '.join(row))
        mylist = t.split(",")

        for element in mylist:                  ##reads each element of each row
            if colNum == 0:                    
                Names[rowNum] = element         ##adds the name of each case to the array 'Names'
            elif colNum != 0:
                Matrix[rowNum][colNum-1] = float(element)     ##adds each value into a matrix in Python
            colNum += 1
        rowNum += 1
    
        i += 1                                  ##these lines end the loop after all elements are read
        if i == num:                            ###and ignore any text below the matrix
            break
    
    BranchNum = 0
    thresh = .007

    for row in range(num-1,0,-1):
        if Branch[row] == -1:
            Branch[row] = BranchNum
            for column in range(0,num):
                if Matrix[row][column] <= thresh and Branch[column] == -1:
                    Branch[column] = BranchNum
            BranchNum += 1

    rowNumbers = [1]*(numberOfRows-1)
    rowNum = offSet
    for index in range(0,BranchNum):
        ind = 0
        for element in Branch: 
            if element == index:
                print(str(index) + " " + str(ind + 1) + " " + Names[ind])
                rowNumbers[ind] = rowNum
                rowNum += 1
            ind += 1
        rowNum += 2 ##add more to create more spaces between branches
        print('  ')

    if rowNumbers[0] == 1:
        count = 0
        for o in rowNumbers:
            rowNumbers[count] += 1
            count += 1
            


    return(rowNumbers, BranchNum)
    
    
def saveExcel(book, Season, State):
    print('Saving...')
    book.save("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\"+State+"\\"+ years+"\\Protein Data" + "-" + State + "-" + Season + ".xlsx")

def findNSites(correctSeq):
    oneAgo = False
    twoAgo = False
    nSites = {0}
    loc = 0
    for l in correctSeq:
        if twoAgo:
            if l == 'S' or l == 'T':
                nSites.add(loc)
            twoAgo = False
        if oneAgo:
            twoAgo = True
            oneAgo = False
        if l == 'N':
            oneAgo = True
        loc += 1

    nSites.remove(0)
    return nSites

def compSDs(book,SDoffSet, offSet,years, State):
    sheet = book["Branches-Pro"]
    newSheet = book["Standard Deviations"]
    Next = True
    colNum = 2
    rowNum = offSet
    row1 = offSet
    totalVal = 0
    nRowNum = SDoffSet

    for i in range(1,551):
        cell = newSheet.cell(row=1, column=18+i, value=i)
        cell.alignment = Alignment(horizontal='center')
    
    
    while(Next):
        el = sheet.cell(row=rowNum, column=colNum)
        totalVal += getVal(el.value)
        if el.value == None:
            if (rowNum-row1) > 1:
                totalVal += 100
                mean = totalVal/(rowNum-row1)
                difsSquared = 0
                for i in range(row1,rowNum):
                    el2 = sheet.cell(row=i, column=colNum)
                    difs = getVal(el2.value) - mean
                    difsSquared += difs*difs
                SD = round(math.sqrt(difsSquared/(rowNum-row1)), 3)
            
                if SD == 0:
                    color = 'FF0000'
                elif SD <= 1:
                    color = 'FFFF00'
                else:
                    color = '008000'
                cell = newSheet.cell(row=nRowNum, column=colNum+1, value=SD)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill('solid', fgColor=color)
            colNum += 1
            totalVal = 0
            if colNum == 568:
                if (rowNum-row1) > 1:
                    newSheet.cell(row=nRowNum, column=1, value='Branch' + ' ' + State + ' ' +  years + ' ' + str(nRowNum-1))
                    newSheet.cell(row=nRowNum, column=2, value=(rowNum-row1))
                    nRowNum += 1
                colNum = 2
                rowNum += 2
                row1 = rowNum
                if sheet.cell(row=rowNum, column=colNum).value == None:
                    Next = False
            rowNum = row1
        else:
            rowNum += 1
    return nRowNum + 1
        
                        
    
def getVal(el):
    if el == 'A':
        return 68.0
    elif el == 'C':
        return 73.3
    elif el == 'D':
        return 19.0
    elif el == 'E':
        return 20.3   
    elif el == 'F':
        return 100.0
    elif el == 'G':
        return 58.4
    elif el == 'H':
        return 30.4
    elif el == 'I':
        return 95.8
    elif el == 'K':
        return 40.3
    elif el == 'L':
        return 95.3
    elif el == 'M':
        return 78.2
    elif el == 'N':
        return 36.3
    elif el == 'P':
        return 75.9
    elif el == 'Q':
        return 37.6
    elif el == 'R':
        return 16.7
    elif el == 'S':
        return 46.6
    elif el == 'T':
        return 54.2
    elif el == 'V':
        return 85.4
    elif el == 'W':
        return 89.8
    elif el == 'Y':
        return 90.0
    elif el == 'Z':
        return 0.0
    else:
        return -100

def allStates():
    book = Workbook()
    Season = input("Season (Ex: 14_15): ")
    States = ['CA', 'FL', 'MA', 'NY', 'TX', 'WI']

    
    sheet1 = book.active
    sheet1.title = "Protein Sequence"
    sheet2 = book.create_sheet("Branches-Pro")
    book.create_sheet("Standard Deviations")

    

    year1 = int(Season[:2])
    year2 = int(Season[3:])
    Season1 = str(year1) + '-' + str(year2)
    Season2 = str(year1) + '_' + str(year2)
    years = '20' + str(year1) + '-20' + str(year2)

    rowNumbers= {1}
    offSet = 2
    SDOffSet = 2
    
    for State in States:
        fastaNuc = open("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\" + State+ "\\"+ years +'\\' + State + Season1 + '-NUC.fasta', "r")
        fastaPro = open("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\" + State+ "\\"+ years +'\\' + State + Season1 + '-PRO.fasta', "r")
        Reader = csv.reader(open("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\" + State+ "\\"+ years +'\\Distance Data-Nuc', newline=''), delimiter=' ', quotechar='|')
        (rowNum, Nucs, Names) = createExcelSheetNuc(Season2, fastaNuc)
        offSetSave = max(rowNumbers) + 2
        (rowNumbers, numBranches) = compBranch(Reader, rowNum, offSetSave)
        (correctSeq, fastaPro) = computeMostCommon(fastaPro, Season2)
        nSites = findNSites(correctSeq)
        (offSet, Xs) = createExcelSheetPro(sheet1, sheet2,Season2, correctSeq, fastaPro, rowNumbers, nSites, offSet)
        SDOffSet = compSDs(book, SDOffSet,offSetSave, years, State)
        offSet += 2
        

        
    print('Saving...')
    book.save("U:\\ResearchData\\rdss_hhaim\\LAB PROJECTS\\Influenza\\H3N2 Analyses\\USA\\Across-States-" + Season+".xlsx")



allStates()
