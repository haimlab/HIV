#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 25 13:42:51 2018

@author: Changze Han

This program calculate the historical distribution of different positions' amino acids
STEPS:
1. Change the Input directory, do not delete r
2. Change the Output directory, do not delete r, When you copy the address, keep in mind there's also a slash at the end.
3. Make a name of the Output file
4. Change the PositionRange: if you want to select a single position: [[88,88]]
                            if you want to select multiple positions: [[88,88],[156,156],[276,276]]
                            if you want to select a position range, like from 88 to 637: [[88,637]]
                            you can also mix them: [[88,88],[156,156],[197,295],[392,637]]
5. Change the YearRange, the format is the same as the PositionRange:
                                single year:[1979,1979]
                                year range: [[1979,2014]]
                                mix:[[1979,2014],[1985,2006],[1994,2000],[2008,2008]]
6. Ignore the output file named 'ignore.xlsx', feel free to delete it
7. run Forest run!
"""


import xlrd
import openpyxl


POS_RANGE = (88, 88), (197, 339)
YEAR_RANGE = (1981, 1983), (1979, 2001), (2004, 2013), (1981, 1981)


def main(ifn, ofn):
    workbook = xlrd.open_workbook(ifn)
    sheet = workbook.sheet_by_index(0)
    nRows = sheet.nrows
    nCols = sheet.ncols

    Position = get_position_list()
    test2 = get_year_range_data(YEAR_RANGE, Position)
    write_file()
    write_file_2()


#get data of row y by index, eg:GetRowIndexData(2)
def get_row_index_data(y):
    row = []
    for col in range (nCols):
        row.append(sheet.cell_value((y-1),col))
    return row

#get data of column x by index
def get_col_index_data(x):
    col = []
    for row in range (nRows):
        col.append(sheet.cell_value(row,(x-1)))
    return col

#get data of column x by title of first row , eg: GetColData('Year')
def get_col_data(x):
    
    for i in get_row_index_data(1):
        if i == x : 
            a = get_col_index_data(get_row_index_data(1).index(i) + 1)
            del a[0]
            return a
#get data at a specific location, eg:GetDataAt(1,2)
def get_data_at(y, x):
    return sheet.cell_value((y-1),(x-1))


#get position List
def get_position_list():
    result = []
    for i in POS_RANGE:
        for j in get_row_index_data(1):
            if (type(j)== int) or (type(j)== float):
                if (j>= i[0]) and (j <= i[1]) :
                    result.append(int(j))
    return result
            
    


#get data of position p during year range y
def get_year_range_data(y, p):
    finalfinal=[]
    for k in p:
        pos = get_row_index_data(1).index(k) + 1
        final = []
        year_range = get_col_data('Year')
        for i in y:
            result = []
            num = 1
            for j in year_range:
                num = num + 1
                if (j >= i[0]) and (j <= i[1]) :
                    result.append(get_data_at(num, pos))
                    
            final.append(result)
        finalfinal.append(str(k))
        finalfinal.append(final)
        
        
    return finalfinal
   
     




#write into excel file

# noinspection PyShadowingNames
def write_file():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'

#Loop to set the value of each cell
    r_count = 1
    c_count = 1
    for i in range(len(test2)):
        #print (i)
        if type(test2[i]) == str:
            sheet.cell(row=r_count, column=c_count).value = test2[i]
            c_count = c_count + 1
        else:
            for j in range(len(test2[i])):
            #print(j)
                for k in range(len(test2[i][j])):
                
                    sheet.cell(row=r_count, column=c_count).value=test2[i][j][k]
                    r_count = r_count + 1
                r_count = 1
                c_count = c_count + 1
    wb.save(Output+"ignore.xlsx")
    



#open the ignore.xlsx, then count the amino acid
file_location2 = Output+"ignore.xlsx"
workbook2 = xlrd.open_workbook(file_location2)
sheet2 = workbook2.sheet_by_index(0)
nRows2 = sheet2.nrows
nCols2 = sheet2.ncols

def get_col_index_data_2(x):
    col = []
    for row in range (nRows2):
        col.append(sheet2.cell_value(row,(x-1)))
    return col

def count():
    name = ['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y','Z']
    final = []
    for col in range(nCols2):
        col_count = []
        if get_col_index_data_2(col + 1)[0].isdigit() :
            final.append(get_col_index_data_2(col + 1)[0])
            final.append(name)
        else:
            for aa in name:      
                num = 0
                for n in get_col_index_data_2(col + 1):
                
                    if aa == n:
                        num = num + 1
                col_count.append(num)
            final.append(col_count)
    return final

test3 = count()
test4 = count()
# add one row which is the total number of amino acids of each position 
def calc_sum():
    for i in range(len(test4)):       
        num = 0
        if type(test4[i]) is str:
            test4[i] = ''
        elif type(test4[i][0]) is str:
                    test4[i] = "Sum"
        else:
            for j in test4[i]:
                num = num + j
            test4[i] = num
    return test4

#write into excel file
# noinspection PyShadowingNames
def write_file_2():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'

#Loop to set the value of each cell
    col_c = 3
    for p in range(len(Position)):
        
        for i in range(len(YEAR_RANGE)):
            sheet.cell(row=1, column=col_c).value=str(YEAR_RANGE[i])
            col_c = col_c+1
        col_c = col_c + 2
    col_c_c = 1
    for i in range(len(test3)):
        
        if type(test3[i]) == str:  
            col_c_c = col_c_c+1
            sheet.cell(row=1, column=col_c_c).value=test3[i]
        else:
            for j in range(len(test3[i])):
                sheet.cell(row=j+2, column=col_c_c).value=test3[i][j]
            col_c_c = col_c_c+1
    for i in range(len(test4)):
        sheet.cell(row=24, column=i+1).value=test4[i]
    wb.save(Output + OutputName)
    







