################################################
# Program to compute the standard deviation of
# influenza strains within branches of a season.
################################################
#Author Kallin Khan, Department of Microbiology
#Updated: 4/3/17
################################################

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl import Workbook
import math

def ComputeStandardDevs(book,SDoffSet, offSet,years, State):
    sheet = book["Branches-Pro"]
    if SDoffSet == 2:
        newSheet = book.create_sheet("Standard Deviations")
    else:
        newSheet = book["Standard Deviations"]
    Next = True
    colNum = 2
    rowNum = offSet
    row1 = offSet
    totalVal = 0
    nRowNum = SDoffSet

    for i in range(-15,551):
        cell = newSheet.cell(row=1, column=18+i, value=i)
        cell.alignment = Alignment(horizontal='center')
    
    
    while(Next):
        el = sheet.cell(row=rowNum, column=colNum)
        totalVal += getVal(el.value)
        if el.value == None:
            if (rowNum-row1) > 1:
                totalVal += 100
                mean = totalVal/(rowNum-row1)
                difsSquared = 0
                for i in range(row1,rowNum):
                    el2 = sheet.cell(row=i, column=colNum)
                    difs = getVal(el2.value) - mean
                    difsSquared += difs*difs
                SD = round(math.sqrt(difsSquared/(rowNum-row1)), 3)
            
                if SD == 0:
                    SD = 0.001
                    color = 'FF0000'
                elif SD <= 1:
                    color = 'FFFF00'
                else:
                    color = '008000'
                logSD = math.log(SD,10)
                cell = newSheet.cell(row=nRowNum, column=colNum+1, value=logSD)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill('solid', fgColor=color)
            colNum += 1
            totalVal = 0
            if colNum == 568:
                if (rowNum-row1) > 1:
                    newSheet.cell(row=nRowNum, column=1, value='Branch' + ' ' + State + ' ' +  years + ' ' + str(nRowNum-1))
                    newSheet.cell(row=nRowNum, column=2, value=(rowNum-row1))
                    nRowNum += 1
                colNum = 2
                rowNum += 2
                row1 = rowNum
                if sheet.cell(row=rowNum, column=colNum).value == None:
                    Next = False
            rowNum = row1
        else:
            rowNum += 1
    return nRowNum + 1
        
def aveSDs(book, end):
    sheet = book["Standard Deviations"]

    for i in range(-15,551):
        cell = sheet.cell(row=end, column=18+i, value=i)
        cell.alignment = Alignment(horizontal='center')

    Next = True
    rowNum = 2
    colNum = 3
    totalVal = 0
    count = 0

    while(Next):
        if colNum == 569:
            break
        elif rowNum == end:
            cell = sheet.cell(row=rowNum+1, column=colNum, value=totalVal/count)
            cell.alignment = Alignment(horizontal='center')
            rowNum = 2
            totalVal = 0
            colNum +=1
            count = 0
            continue
        el = sheet.cell(row=rowNum, column=colNum).value
        if el == None:
            rowNum += 1
            continue
        totalVal += el
        rowNum += 1
        count += 1                        
    
def getVal(el):
    if el == 'A':
        return 68.0
    elif el == 'C':
        return 73.3
    elif el == 'D':
        return 19.0
    elif el == 'E':
        return 20.3   
    elif el == 'F':
        return 100.0
    elif el == 'G':
        return 58.4
    elif el == 'H':
        return 30.4
    elif el == 'I':
        return 95.8
    elif el == 'K':
        return 40.3
    elif el == 'L':
        return 95.3
    elif el == 'M':
        return 78.2
    elif el == 'N':
        return 36.3
    elif el == 'P':
        return 75.9
    elif el == 'Q':
        return 37.6
    elif el == 'R':
        return 16.7
    elif el == 'S':
        return 46.6
    elif el == 'T':
        return 54.2
    elif el == 'V':
        return 85.4
    elif el == 'W':
        return 89.8
    elif el == 'Y':
        return 90.0
    elif el == 'Z':
        return 0.0
    elif el == '-':
        return 150.0
    else:
        print("UNLISTED VALUE: " + el)
        return -100
